import watson_developer_cloud
import os
import codecs
from watson_developer_cloud.personality_insights_v3 import Profile
import re
import json
import csv   
from openpyxl import load_workbook

consumptionPreferences=None  #can take value True or None

#load Excel workbook
wb = load_workbook(filename = 'Cognitive Assessment January 2018_data.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
personsDataList= []

#Read headers from the input Excel sheet
headers=[sheet['A' + str(1)].value, sheet['B' + str(1)].value, sheet['C' + str(1)].value, sheet['D' + str(1)].value, sheet['E' + str(1)].value, sheet['F' + str(1)].value, sheet['G' + str(1)].value]

personData = {}  
personData['person'] = []  # will contain final output json file data
#Read each persons data and put them in a list
for row in range(2, sheet.max_row + 1):
    tempList=[]
    
    tempList.append(sheet['A' + str(row)].value)
    tempList.append(sheet['B' + str(row)].value)
    tempList.append(sheet['C' + str(row)].value)
    tempList.append(str(sheet['D' + str(row)].value.year) + '-' + str(sheet['D' + str(row)].value.month) + '-' + str(sheet['D' + str(row)].value.day)) # convert the date time to string
    tempList.append(sheet['E' + str(row)].value)
    tempList.append(sheet['F' + str(row)].value)
    tempList.append(sheet['G' + str(row)].value)
    personsDataList.append(tempList)
    
for person in personsDataList:
    emailsTextFile=open("profile.txt", "w", encoding="utf-8") 
    emailsTextFile.write(re.sub('[^A-Za-z]+', ' ', person[1])) #remove special characters and digits from emails
    emailsTextFile.close()
    
    #use Watson PersonalityInsightsV3
    personality_insights = watson_developer_cloud.PersonalityInsightsV3(version='2017-10-13',username='94cb29d7-6094-a2ad57f',password='2ZiLJ')

    with open(os.path.join(os.path.dirname('__file__'), 'profile.txt')) as personality_text:
        response = personality_insights.profile( personality_text, content_type='text/plain;charset=utf-8', content_language='en', accept_language='en',raw_scores=True, consumption_preferences=consumptionPreferences)
   
    modifiedJsonOutput={}
    count=0;
    
    # Write Excel file data for the output json file
    for column in person:
        modifiedJsonOutput[headers[count]]=column
        count=count+1
    
    # Write the output returned by the watson getProfile API
    modifiedJsonOutput['word_count']=response['word_count']
    modifiedJsonOutput['processed_language']=response['processed_language']
    modifiedJsonOutput['personality']=response['personality']
    modifiedJsonOutput['needs']=response['needs']
    modifiedJsonOutput['values']=response['values']
    
    if consumptionPreferences==True:
        modifiedJsonOutput['consumption_preferences']=response['consumption_preferences']
    
    modifiedJsonOutput['warnings']=response['warnings']
    personData['person'].append(modifiedJsonOutput)
                          
# Write final json output. Json file combines the input excel file and the output returned by the watson getProfile API
finalJsonOutput=json.dumps(personData,ensure_ascii=False,indent=2)
finalOutputFile=open("output.json", "w", encoding="utf-8")
finalOutputFile.write(finalJsonOutput)
finalOutputFile.close()

#Now store only the key results in a csv file for analysis purpose
csvOutput = []
excelFileHeaders=[]
csvHeaderRow=[]
#now copy the headers from the json file
for name in headers:
    if name!='Emails': # Copy excel file headers. Emails not copied as it will not be written in the CSV
        excelFileHeaders.append(name)
        csvHeaderRow.append(name)

personDataList=personData['person']
firstPersonPersonality=personDataList[0]['personality']  # copy the personality related watson getProfile API headers (from the first person only)

for bigFiveTrait in firstPersonPersonality:
    csvHeaderRow.append(bigFiveTrait['name'] + '-RawScore') # copy raw score
    csvHeaderRow.append(bigFiveTrait['name'] + '-Percentile') # copy percentile
csvOutput.append(csvHeaderRow) # copy the header row
    
for person in personDataList:
    csvRow=[]  # create a csv file row for each person
    
    for headerName in excelFileHeaders:  #copy excel file info for each person
        csvRow.append(person[headerName])
        
    for bigFiveTrait in person['personality']: #copy personality related info for each person
        csvRow.append(bigFiveTrait['raw_score'])
        csvRow.append(bigFiveTrait['percentile'])
    
    csvOutput.append(csvRow)

#Now store only the key and key results in a csv file for analysis purpose
csvWriterObj=open('csvoutput.csv', 'w')
writer = csv.writer(csvWriterObj)
for row in csvOutput:
    writer.writerow(row)
csvWriterObj.close()     

print("Done")
